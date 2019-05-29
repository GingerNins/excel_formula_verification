import openpyxl
from openpyxl import Workbook, utils
from variable import Formula, Name, Variable


def process_template_file(filename: str) -> dict:
    """
    Processes the Excel file aggregating all formulas, named ranges, constants
    matching them together, storing the list of formulas and list of names in a dict
    :param filename: name of template file
    :return: dict of formula list and named ranges list
    """
    try:
        wb: Workbook = openpyxl.load_workbook(filename)
    except PermissionError as e:
        print(e)
        exit(1)

    named_ranges: list = _get_named_ranges(wb)
    formulas, constants = _get_formulas_and_constants(wb)

    variables = {'formulas': formulas, 'names': named_ranges, 'constants': constants}

    _match_output_data(filename, variables)

    wb.close()

    return variables


def _match_output_data(filename, items):
    """
    Matches the items to the the data values in the Excel file
    :param filename: filename to open as data_only
    :param items: items to match
    :return: n/a
    """
    try:
        wb_data: Workbook = openpyxl.load_workbook(filename, data_only=True)
    except PermissionError as e:
        print(e)
        exit(1)

    constants = items['constants']
    formulas = items['formulas']
    named_ranges = items['names']

    for c in constants:
        c.set_name(named_ranges)
        c.set_output(wb_data)

    for f in formulas:
        f.set_name(named_ranges)
        f.set_output(wb_data)
        f.update_variables(named_ranges)

    for n in named_ranges:
        n.set_is_used(formulas)
        n.set_output(wb_data)

    wb_data.close()


def _get_named_ranges(wb) -> list:
    """
    Aggregates all named ranges into a list
    :return: list of all named ranges
    """
    named_range_list = []

    for dn in wb.defined_names.definedName:
        name = dn.name
        scope = wb.sheetnames[dn.localSheetId] if dn.localSheetId is not None else 'Workbook'
        dest = wb.defined_names.get(name, dn.localSheetId).destinations

        """
        Destinations usually return 0 (global scope) or 1 (Workbook scope) item(s) unless there is a named
        range that exists on multiple sheets with the same name having worksheet level scope
        """
        for sheet_name, rng in dest:
            ws = wb[sheet_name]
            # If the range is only a single cell
            if not isinstance(ws[rng], tuple):
                named_range_list.append(
                    Name(sheet=sheet_name, name=name, scope=scope, cell=ws[rng]))
                break
            # If the named range contains multiple cells
            else:
                for c in ws[rng]:
                    named_range_list.append(
                        Name(sheet=sheet_name, name=name, scope=scope, cell=c[0]))
                break
        else:
            # Global Constants
            named_range_list.append(
                Name(name=name, scope=scope, value=dn.attr_text, is_global=True))

    return named_range_list


def _get_formulas_and_constants(wb) -> tuple:
    """
    Aggregates all formulas and constants in the workbook in a list.
    Constants are cells that contain values necessary for calculations but are
    not formulas themselves.  These are usually manually entered nuumbers by
    the users.
    :return: a tuple of a list of Formulas and list of Constants
    """
    formula_list = []
    constants_list = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        last_cell = utils.get_column_letter(ws.max_column) + str(ws.max_row)
        working_range = ws['A1':last_cell]

        for row in working_range:
            for cell in row:
                # Skip the formula if:
                # - It's blank
                # - It's a string and doesn't start with "=" (indicates just text)
                if cell.value is None or (isinstance(cell.value, str) and not cell.value.startswith('=')):
                    continue

                # FIXME: remove the '=' thing?
                # Note: If i don't check, then I won't know if Variable or Formula
                # Note: Should I perform the check here? or change the variable class
                #   - so that all formulas/variables/names? are the same object
                #   - with different flag types?
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    value = cell.value[1:]
                    formula_list.append(Formula(sheet=sheet_name, cell=cell, value=value))
                else:
                    value = cell.value
                    constants_list.append(Variable(sheet=sheet_name, cell=cell, value=value))

    return formula_list, constants_list



